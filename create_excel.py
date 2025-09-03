from helper_functions import (
    get_sql_dataframe,
    get_size_order,
    get_brand_image,
    resize_image,
)
from pandas import ExcelWriter, DataFrame
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
import sys

def create_xlsx_orderform(
    brand: str,
    season: str,
    category: list,
    currency: str,
    use_size: str,
    colour_scheme: str,
    password: str,
    file_path: str,
    file_path_error: str,
) -> None:

    if getattr(sys, "frozen", False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")

    with open(
        os.path.join(base_path, "t-sql", "style query.sql"),
        "r",
    ) as file:
        style_sql = file.read()

    replacements = {
        "{brand}": brand,
        "{season}": season,
        "{category}": category,
        "{currency}": currency,
    }

    for key, value in replacements.items():
        style_sql = style_sql.replace(key, value)

    df_style, error = get_sql_dataframe(style_sql)

    if error:
        print(str(error))
        return error

    df_style = df_style.drop("Row No", axis=1)

    with open(
        os.path.join(base_path, "t-sql", "sku query.sql"),
        "r",
    ) as file:
        sku_sql = file.read()

    for key, value in replacements.items():
        sku_sql = sku_sql.replace(key, value)

    df_sku, error = get_sql_dataframe(sku_sql)

    if error:
        print(str(error))
        return error

    # Add in the invalid sku data

    with open(
        os.path.join(base_path, "t-sql", "Invalid sku query.sql"),
        "r",
    ) as file:
        invalid_sku_sql = file.read()
        
    for key, value in replacements.items():
        invalid_sku_sql = invalid_sku_sql.replace(key, value)

    df_invalid_sku, error = get_sql_dataframe(invalid_sku_sql)

    if error:
        print(str(error))
        return error

    if use_size == "Size 1":
        df_sku = df_sku.drop(["EU Size", "US Size"], axis=1)
        df_sku = df_sku.rename(columns={"Size 1": "Size"})

    elif use_size == "EU":
        df_sku = df_sku.drop(["Size 1", "US Size"], axis=1)
        df_sku = df_sku.rename(columns={"EU Size": "Size"})

    elif use_size == "US":
        df_sku = df_sku.drop(["Size 1", "EU Size"], axis=1)
        df_sku = df_sku.rename(columns={"US Size": "Size"})

    df_duplicates = df_sku[
        df_sku.duplicated(subset=["Preorder Code", "Style No", "Size"], keep=False)
    ]

    df_sku = df_sku.drop_duplicates(subset=["Style No", "Size"], keep="first")

    df_style_error = df_style[df_style["Size Cat"].isnull()]
    df_sku_error = df_sku[df_sku["Size Cat"].isnull()]

    if len(df_sku_error) or len(df_duplicates) > 0 or len(df_invalid_sku) > 0:

        df_sku = df_sku[~df_sku["Style No"].isin(df_sku_error["Style No"])]

        df_sku = df_sku[~df_sku["Style No"].isin(df_duplicates["Style No"])]

        df_style = df_style[~df_style["Style No"].isin(df_sku_error["Style No"])]

        df_style = df_style[~df_style["Style No"].isin(df_duplicates["Style No"])]

        with ExcelWriter(file_path_error, engine="openpyxl") as writer:
            df_style_error.to_excel(writer, sheet_name="Style Error", index=False)
            df_sku_error.to_excel(writer, sheet_name="SKU Error", index=False)
            df_duplicates.to_excel(writer, sheet_name="Duplicates", index=False)
            df_invalid_sku.to_excel(writer, sheet_name="Invalid SKU", index=False)

        # The below code sets the formatting colours for each orderform workbook

    if colour_scheme == "Shiner Ltd":
        main_colour = "FF2371B6"
        alt_colour = "FF003DAB"
        light_colour = "FFD1EFFA"
    elif colour_scheme == "Shiner B.V":
        main_colour = "FFFCA800"
        alt_colour = "FFF96F00"
        light_colour = "FFFCEDD6"
    elif colour_scheme == "Shiner LLC":
        main_colour = "FF2DAB66"
        alt_colour = "FF087329"
        light_colour = "FFE2EFDA"

    sheet1_name = f"{brand} {season} {currency}"
    if len(sheet1_name) > 31:
        sheet1_name = sheet1_name[:31]

    with ExcelWriter(file_path, engine="openpyxl") as writer:
        df_style.to_excel(writer, sheet_name=sheet1_name, index=False)
        df_sku.to_excel(writer, sheet_name="SKU Data", index=False)
        DataFrame().to_excel(writer, sheet_name="Data Import", index=False)

        workbook = writer.book

        sheet1 = workbook.worksheets[0]
        sheet2 = workbook.worksheets[1]
        sheet3 = workbook.worksheets[2]

        no_border = Border(
            left=Side(border_style=None),
            right=Side(border_style=None),
            top=Side(border_style=None),
            bottom=Side(border_style=None),
        )

        for cell in sheet1[1][: df_style.shape[1]]:
            cell.font = Font(bold=False)
            cell.border = no_border

        for cell in sheet2[1][: df_sku.shape[1]]:
            cell.font = Font(bold=False, color="FFFFFF")
            cell.border = no_border

        if currency == "GBP":
            sheet1["H1"] = "£ Price"
            sheet1["I1"] = "£ SRP"
            sheet1["J1"] = "Qty Total"
            sheet1["K1"] = "£ Total"

            sheet2["H1"] = "£ Price"
            sheet2["I1"] = "£ SRP"

        elif currency == "EUR":
            sheet1["H1"] = "€ Price"
            sheet1["I1"] = "€ SRP"
            sheet1["J1"] = "Qty Total"
            sheet1["K1"] = "€ Total"

            sheet2["H1"] = "€ Price"
            sheet2["I1"] = "€ SRP"

        elif currency == "USD":
            sheet1["H1"] = "$ Price"
            sheet1["I1"] = "$ SRP"
            sheet1["J1"] = "Qty Total"
            sheet1["K1"] = "£ Total"

            sheet2["H1"] = "$ Price"
            sheet2["I1"] = "$ SRP"

        sheet2["K1"] = "Quantity"

        size_cat_count = df_style["Size Cat"].nunique()

        if size_cat_count > 1:
            sheet1.insert_rows(idx=2, amount=size_cat_count - 1)

        sheet1.insert_rows(idx=1, amount=8)

        size_counts = df_sku.groupby("Size Cat")["Size"].nunique()

        max_size_count = size_counts.max()

        sheet1.insert_cols(idx=1, amount=1)
        sheet1.insert_cols(idx=9, amount=max_size_count + 1)

        unique_size_cats = df_style["Size Cat"].unique()

        Size_cat_order = [
            "Misc",
            "MT",
            "L",
            "FT",
            "FLOZ",
            "M",
            "CM",
            "MM",
            "IN",
            "TODD",
            "JNR",
            "CHILD",
            "Yth SWS",
            "Yth Heelys",
            "Yth Footwear",
            "Youth",
            "Youth Misc",
            "Women Misc",
            "Women",
            "Accessory",
            "SWS",
            "Heelys",
            "Footwear",
            "Pants",
            "Waist",
            "Adult Misc",
            "Adult /",
            "Adult",
        ]

        size_cat_dict = {name: idx for idx, name in enumerate(Size_cat_order)}

        in_custom_order = [cat for cat in unique_size_cats if cat in size_cat_dict]
        not_in_custom_order = sorted(
            [cat for cat in unique_size_cats if cat not in size_cat_dict]
        )

        in_custom_order_sorted = sorted(
            in_custom_order, key=lambda cat: size_cat_dict[cat]
        )

        sorted_size_cats = not_in_custom_order + in_custom_order_sorted

        for idx, size_cat in enumerate(sorted_size_cats, start=9):
            sheet1.cell(row=idx, column=9, value=size_cat)

        size_row = 9
        input_qty_row = size_row + size_cat_count
        input_qty_end = len(df_style) + size_cat_count + 8
        start_col = 10

        for size_cat in sorted_size_cats:
            sorted_size_list = get_size_order(size_cat, use_size)
            unique_sizes = list(df_sku[df_sku["Size Cat"] == size_cat]["Size"].unique())

            in_sorted_list = [size for size in unique_sizes if size in sorted_size_list]
            not_in_sorted_list = [
                size for size in unique_sizes if size not in sorted_size_list
            ]

            in_sorted_list.sort(key=lambda x: sorted_size_list.index(x))

            not_in_sorted_list.sort()

            final_sorted_sizes = in_sorted_list + not_in_sorted_list

            for size in final_sorted_sizes:
                sheet1.cell(row=size_row, column=start_col, value=size)

                for current_row in range(input_qty_row, input_qty_end + 1):
                    
                    style_no = sheet1.cell(row=current_row, column=2).value
                    size_cat_desc = sheet1.cell(row=current_row, column=5).value

                    current_cell = sheet1.cell(row=current_row, column=start_col)

                    current_cell.number_format = "#,##0"

                    if current_cell.value is None or current_cell.value < 1:
                        exists_count = len(
                            df_sku.query(
                                "`Style No` == @style_no and `Size Cat` == @size_cat_desc and Size == @size"
                            )
                        )
                        current_cell.value = exists_count

                    if size == None:
                        current_cell.value = 0

                start_col += 1

            size_row += 1
            start_col = 10

        for row in range(input_qty_row, input_qty_end + 1):

            size_cat_ws = sheet1.cell(row=row, column=5).value

            number_of_size_cat_sizes = df_sku[df_sku["Size Cat"] == size_cat_ws][
                "Size"
            ].nunique()

            for col in range(10, 10 + max_size_count):

                col_counter = col - 9

                if number_of_size_cat_sizes < col_counter:

                    sheet1.cell(row=row, column=col).value = 0

        # The below code starts to format the orderform worksheet

        for header_col in range(2, 9):
            start_cell = sheet1.cell(row=9, column=header_col)
            end_cell = sheet1.cell(row=input_qty_row - 1, column=header_col)

            sheet1.merge_cells(
                start_row=start_cell.row,
                start_column=start_cell.column,
                end_row=end_cell.row,
                end_column=end_cell.column,
            )

        price_col_Start = 10 + max_size_count
        price_col_end = price_col_Start + 4

        for header_col in range(price_col_Start, price_col_end):
            start_cell = sheet1.cell(row=9, column=header_col)
            end_cell = sheet1.cell(row=input_qty_row - 1, column=header_col)

            sheet1.merge_cells(
                start_row=start_cell.row,
                start_column=start_cell.column,
                end_row=end_cell.row,
                end_column=end_cell.column,
            )

        header_fill = PatternFill(
            start_color=main_colour, end_color=main_colour, fill_type="solid"
        )
        light_fill = PatternFill(
            start_color=light_colour, end_color=light_colour, fill_type="solid"
        )
        header_alignment = Alignment(vertical="center", horizontal="center")

        start_cell = sheet1.cell(row=9, column=2).coordinate
        end_cell = sheet1.cell(
            row=8 + size_cat_count, column=13 + max_size_count
        ).coordinate

        cell_range = f"{start_cell}:{end_cell}"

        for row in sheet1[cell_range]:
            for cell in row:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = header_alignment

        medium_border = Side(border_style="medium", color=alt_colour)
        dashed_border = Side(border_style="dashed", color=alt_colour)
        left_border = Border(
            top=medium_border, left=medium_border, bottom=medium_border
        )
        right_border = Border(
            top=medium_border, right=medium_border, bottom=medium_border
        )

        start_col = 2
        end_col = 14 + max_size_count

        if size_cat_count == 1:
            for col in range(start_col, end_col):
                sheet1.cell(row=9, column=col).border = Border(
                    top=medium_border, bottom=medium_border
                )

        else:

            for col in range(start_col, end_col):
                sheet1.cell(row=9, column=col).border = Border(top=medium_border)
                sheet1.cell(row=8 + size_cat_count, column=col).border = Border(
                    bottom=medium_border
                )

        for row in range(9, 9 + size_cat_count):
            sheet1.cell(row=row, column=2).border = left_border
            sheet1.cell(row=row, column=13 + max_size_count).border = right_border

        thin_border = Side(border_style="thin", color=alt_colour)
        left_grid_border = Border(left=medium_border, bottom=thin_border)
        right_grid_border = Border(right=medium_border, bottom=thin_border)

        for row in range(input_qty_row, input_qty_end + 1):

            for col in range(start_col, end_col):
                sheet1.cell(row=row, column=col).border = Border(bottom=thin_border)

        size_border_grid = Border(
            left=thin_border, right=thin_border, bottom=thin_border
        )
        thin_side_border = Border(left=thin_border, right=thin_border)

        left_dashed_border = Border(
            left=thin_border, right=dashed_border, bottom=thin_border
        )
        middle_dashed_border = Border(
            left=dashed_border, right=dashed_border, bottom=thin_border
        )
        right_dashed_border = Border(
            left=dashed_border, right=medium_border, bottom=thin_border
        )

        if currency == "GBP":
            currency_format = '_-£* #,##0.00_-;-£* #,##0.00_-;_-£* "-"??_-;_-@_-'
        elif currency == "EUR":
            currency_format = '_-€* #,##0.00_-;-€* #,##0.00_-;_-€* "-"??_-;_-@_-'
        elif currency == "USD":
            currency_format = '_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-'

        for row in range(input_qty_row, input_qty_end + 1):

            if sheet1.cell(row=row, column=3).value == "C/O NC":
                sheet1.cell(row=row, column=3).fill = header_fill
                sheet1.cell(row=row, column=3).font = Font(color="FFFFFF", bold=False)

            elif sheet1.cell(row=row, column=3).value == "C/O":
                sheet1.cell(row=row, column=3).fill = light_fill
                sheet1.cell(row=row, column=3).font = Font(color=alt_colour, bold=False)

            sheet1.cell(row=row, column=9).fill = PatternFill(
                start_color="FF000000", end_color="FF000000", fill_type="solid"
            )

            sheet1.cell(row=row, column=10 + max_size_count).number_format = (
                currency_format
            )
            sheet1.cell(row=row, column=11 + max_size_count).number_format = (
                currency_format
            )
            sheet1.cell(row=row, column=13 + max_size_count).fill = light_fill
            sheet1.cell(row=row, column=13 + max_size_count).font = Font(
                color=alt_colour, bold=False
            )

            sheet1.cell(
                row=row,
                column=12 + max_size_count,
                value=f"=SUM({get_column_letter(10)}{row}:{get_column_letter(9 + max_size_count)}{row})",
            )
            sheet1.cell(row=row, column=12 + max_size_count).number_format = "#,##0"
            sheet1.cell(
                row=row,
                column=13 + max_size_count,
                value=f"={get_column_letter(10 + max_size_count)}{row} * {get_column_letter(12 + max_size_count)}{row}",
            )
            sheet1.cell(row=row, column=13 + max_size_count).number_format = (
                currency_format
            )

            sheet1.cell(row=row, column=9).border = thin_side_border
            sheet1.cell(row=row, column=10 + max_size_count).border = left_dashed_border
            sheet1.cell(row=row, column=12 + max_size_count).border = (
                middle_dashed_border
            )
            sheet1.cell(row=row, column=13 + max_size_count).border = (
                right_dashed_border
            )

            for col in range(10, 10 + max_size_count):
                sheet1.cell(row=row, column=col).border = size_border_grid

        for row in range(input_qty_row, input_qty_end + 1):

            sheet1.cell(row=row, column=start_col).border = left_grid_border
            sheet1.cell(row=row, column=end_col - 1).border = right_grid_border

        for col in range(start_col, end_col):
            sheet1.cell(row=input_qty_end + 1, column=col).border = Border(
                top=medium_border
            )

        # set orderform totals

        sheet1.cell(
            input_qty_end + 1,
            12 + max_size_count,
            value=f"=SUM({get_column_letter(12 + max_size_count)}{input_qty_row}:{get_column_letter(12 + max_size_count)}{input_qty_end})",
        )
        sheet1.cell(input_qty_end + 1, 12 + max_size_count).font = Font(
            color=alt_colour, bold=True
        )
        sheet1.cell(input_qty_end + 1, 13 + max_size_count).number_format = "#,##0"

        sheet1.cell(
            input_qty_end + 1,
            13 + max_size_count,
            value=f"=SUM({get_column_letter(13 + max_size_count)}{input_qty_row}:{get_column_letter(13 + max_size_count)}{input_qty_end})",
        )
        sheet1.cell(input_qty_end + 1, 13 + max_size_count).font = Font(
            color=alt_colour, bold=True
        )
        sheet1.cell(input_qty_end + 1, 13 + max_size_count).number_format = (
            currency_format
        )

        every_thin_border = Border(
            top=thin_border, bottom=thin_border, left=thin_border, right=thin_border
        )
        every_thin_border_medium_top = Border(
            top=medium_border, bottom=thin_border, left=thin_border, right=thin_border
        )
        every_border_blank_medium_top = Border(top=medium_border)

        for row in range(input_qty_row, input_qty_end + 1):
            for col in range(10, 10 + max_size_count):
                if sheet1.cell(row=row, column=col).value == 1:
                    sheet1.cell(row=row, column=col).protection = Protection(
                        locked=False
                    )
                    sheet1.cell(row=row, column=col, value="")

                    if row == input_qty_row:
                        sheet1.cell(row=row, column=col).border = (
                            every_thin_border_medium_top
                        )
                    else:
                        sheet1.cell(row=row, column=col).border = every_thin_border

                elif sheet1.cell(row=row, column=col).value == 0:
                    sheet1.cell(row=row, column=col).protection = Protection(
                        locked=True
                    )
                    sheet1.cell(row=row, column=col, value="")

                    sheet1.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid"
                    )

                    if row == input_qty_row:
                        sheet1.cell(row=row, column=col).border = (
                            every_border_blank_medium_top
                        )
                    else:
                        sheet1.cell(row=row, column=col).border = None

        sheet1.column_dimensions["E"].hidden = True

        sheet1_widths = {
            "A": 2.78,
            "B": 14.78,
            "C": 7.49,
            "D": 5.35,
            "F": 27.89,
            "G": 27.89,
            "H": 20.78,
            "I": 15.78,
        }

        for col, width in sheet1_widths.items():
            sheet1.column_dimensions[col].width = width

        for col in range(10, 10 + max_size_count):
            col_letter = get_column_letter(col)
            sheet1.column_dimensions[col_letter].width = 6.78

        for col in range(11 + max_size_count, 13 + max_size_count):
            col_letter = get_column_letter(col)
            sheet1.column_dimensions[col_letter].width = 10.78

        sheet1.column_dimensions[get_column_letter(13 + max_size_count)].width = 12.78

        for col in range(6, 8):
            sheet1.cell(row=2, column=col).border = Border(top=medium_border)
            sheet1.cell(row=7, column=col).border = Border(bottom=medium_border)

        for row in range(2, 8):
            sheet1.merge_cells(
                start_row=row,
                start_column=6,
                end_row=row,
                end_column=7,
            )
            if row == 2:
                sheet1.cell(row=row, column=6).border = Border(
                    left=medium_border, top=medium_border
                )
                sheet1.cell(row=row, column=7).border = Border(
                    right=medium_border, top=medium_border
                )
            elif row == 7:
                sheet1.cell(row=row, column=6).border = Border(
                    left=medium_border, bottom=medium_border
                )
                sheet1.cell(row=row, column=7).border = Border(
                    right=medium_border, bottom=medium_border
                )
            else:
                sheet1.cell(row=row, column=6).border = Border(left=medium_border)
                sheet1.cell(row=row, column=7).border = Border(right=medium_border)

        for col in range(6, 8):
            for row in range(2, 8):
                cell = sheet1.cell(row=row, column=col)
                cell.fill = light_fill
                cell.font = Font(color=alt_colour, bold=False)
                sheet1.cell(row=row, column=col).protection = Protection(locked=False)

        sheet1["F2"] = "Brand: " + brand
        sheet1["F3"] = "Season: " + season
        sheet1["F4"] = "Customer Name: "
        sheet1["F5"] = "Drop 1 ETA: "
        sheet1["F6"] = "Drop 2 ETA: "
        sheet1["F7"] = "Deadline: "

        image_path = os.path.join(base_path, "logos", str(get_brand_image(brand)))
        image = Image(image_path)

        image = resize_image(image, new_width=120)

        sheet1.add_image(image, "B2")

        sku_data_rows = len(df_sku) + 1
        table_range = f"A1:K{sku_data_rows}"

        sku_data_table = Table(displayName="SKU_Data", ref=table_range)

        style = TableStyleInfo(
            name="TableStyleLight8",
            showFirstColumn=True,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        sku_data_table.tableStyleInfo = style

        sheet2.add_table(sku_data_table)

        formula_sheet_name = "'" + sheet1_name + "'!"
        formula_end_col = get_column_letter(9 + max_size_count)

        for row in range(2, sku_data_rows + 1):
            sku_formula = (
                f"=_xlfn.XLOOKUP($B{row},{formula_sheet_name}$B${input_qty_row}:$B${input_qty_end},"
                f"_xlfn.XLOOKUP($G{row},_xlfn.FILTER({formula_sheet_name}$J$9:${formula_end_col}${8 + size_cat_count},"
                f"{formula_sheet_name}$I$9:$I${8 + size_cat_count} = $C{row}),{formula_sheet_name}$J${input_qty_row}:${formula_end_col}${input_qty_end}))"
            )
            sheet2.cell(row=row, column=11).value = sku_formula

        sheet2_widths = {
            "A": 20.92,
            "B": 14.78,
            "C": 10.78,
            "D": 27.89,
            "E": 27.89,
            "F": 20.78,
            "G": 7.78,
            "H": 8.34,
            "I": 8.34,
            "J": 14.78,
            "K": 10.78,
        }

        for col, width in sheet2_widths.items():
            sheet2.column_dimensions[col].width = width

        preorder_code_list = list(df_sku["Preorder Code"].unique())

        preorder_code_string = ",".join(preorder_code_list)

        dv = DataValidation(type="list", formula1=f'"{preorder_code_string}"')

        sheet3.add_data_validation(dv)

        sheet3["L4"] = preorder_code_list[0]

        dv.add(sheet3["L4"])

        sheet3["K4"] = "Preorder:"

        sheet3["B2"] = (
            "Remove @ from the start of the formulas in B9 & G9! Otherwise the formulas won't spill!"
        )
        sheet3["B2"].font = Font(color=alt_colour, bold=True, size=18)

        sheet3["B4"] = "All Preorder Codes"

        sheet3["B6"] = "Total"
        sheet3["C6"] = f"=SUM('SKU Data'!$K$2:$K${sku_data_rows})"

        sheet3["B8"] = "Item No"
        sheet3["C8"] = "Quantity"

        sheet3["B9"] = (
            f"=_xlfn.FILTER('SKU Data'!$J$2:$K${sku_data_rows},'SKU Data'!$K$2:$K${sku_data_rows} > 0,\"No Qtys Added\")"
        )

        sheet3["G4"] = "Individual Preorder Codes"

        sheet3["G6"] = "Total"
        sheet3["H6"] = (
            f"=SUMIFS('SKU Data'!$K$2:$K${sku_data_rows}, 'SKU Data'!$A$2:$A${sku_data_rows}, $L$4)"
        )

        sheet3["G8"] = "Item No"
        sheet3["H8"] = "Quantity"

        sheet3["G9"] = (
            f"=_xlfn.FILTER('SKU Data'!$J$2:$K${sku_data_rows},('SKU Data'!$K$2:$K${sku_data_rows} > 0) * ('SKU Data'!$A$2:$A${sku_data_rows} = $L$4),\"No Qtys Added\")"
        )

        for row in sheet3["B3:N8"]:
            for cell in row:
                cell.font = Font(color=alt_colour, bold=True)

        sheet3_widths = {
            "A": 2.78,
            "B": 13.78,
            "C": 10.78,
            "G": 13.78,
            "H": 10.78,
            "L": 20.78,
        }

        for col, width in sheet3_widths.items():
            sheet3.column_dimensions[col].width = width

        sheet1.sheet_view.showGridLines = False
        sheet1.protection.set_password(password)

        sheet2.sheet_view.showGridLines = False
        sheet2.protection.set_password(password)
        sheet2.sheet_state = "hidden"

        sheet3.sheet_view.showGridLines = False
        sheet3.sheet_state = "hidden"

    return None